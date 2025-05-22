from django.shortcuts import get_object_or_404, render, redirect
from rest_framework import viewsets
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Vehiculo
from app.utils import generar_alertas
from .models import  Vehiculo, Conductor, Mantenimiento, AlertaMantenimiento, ServicioMantenimiento, DetalleMantenimiento, Asignacion
from .serializers import *
from .forms import AlertaForm, VehiculoForm, ConductorForm, MantenimientoForm 
from django.core.paginator import Paginator

from django.db.models import Q
from django.contrib.auth.decorators import login_required as loginrequired
# Vistas para modelos principales
@loginrequired
def index(request):
    return render(request, "flotas/index.html")

#VISTAS DE VEHICULOS
def listar_vehiculos(request):
    query = request.GET.get("q", "")
    
    if query:
        vehiculos_list = Vehiculo.objects.filter(
            Q(placa__icontains=query) |
            Q(marca__icontains=query) |
            Q(modelo__icontains=query)
        ).order_by('placa')
    else:
        vehiculos_list = Vehiculo.objects.all().order_by('placa')
    
    paginator = Paginator(vehiculos_list, 5)
    page_number = request.GET.get('page')
    vehiculos = paginator.get_page(page_number)

    return render(request, 'flotas/vehiculos/listar_vehiculos.html', {
        'vehiculos': vehiculos,
        'query': query
    })

#CREAR VEHICULO
def crear_vehiculo(request):
    if request.method == 'POST':
        form = VehiculoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_vehiculos')
    else:
        form = VehiculoForm()
    return render(request, 'flotas/vehiculos/crear_vehiculo.html', {'form': form})

# EDITAR VEHICULO
def editar_vehiculo(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)
    if request.method == 'POST':
        form = VehiculoForm(request.POST, instance=vehiculo)
        if form.is_valid():
            form.save()
            return redirect('listar_vehiculos')  # o donde quieras redirigir
    else:
        form = VehiculoForm(instance=vehiculo)
    return render(request, 'flotas/vehiculos/editar_vehiculo.html', {'form': form})


# ELIMINAR VEHICULO
def eliminar_vehiculo(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)
    if request.method == 'POST':
        vehiculo.delete()
        return redirect('listar_vehiculos')
    return render(request, 'flotas/vehiculos/eliminar_vehiculo.html', {'vehiculo': vehiculo})


#VISTAS DE CONDUCTORES
def listar_conductores(request):
    conductores_list = Conductor.objects.all()
    paginator = Paginator(conductores_list, 5) 

    page_number = request.GET.get('page')
    conductores = paginator.get_page(page_number) 

    return render(request, 'flotas/conductores/listar_conductores.html', {'conductores': conductores})


def crear_conductor(request):
    if request.method == 'POST':
        form = ConductorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_conductores')
    else:
        form = ConductorForm()
    return render(request, 'flotas/conductores/crear_conductor.html', {'form': form})


def editar_conductor(request, id):
    conductor = get_object_or_404(Conductor, pk=id)
    if request.method == 'POST':
        form = ConductorForm(request.POST, instance=conductor)
        if form.is_valid():
            form.save()
            return redirect('listar_conductores')
    else:
        form = ConductorForm(instance=conductor)
    return render(request, 'flotas/conductores/editar_conductor.html', {'form': form})


def eliminar_conductor(request, id):
    conductor = get_object_or_404(Conductor, id=id)
    conductor.delete()
    return redirect('listar_conductores')

#VISTAS DE MANTENIMIENTOS
def listar_mantenimientos(request):
    mantenimientos_list = Mantenimiento.objects.all()
    paginator = Paginator(mantenimientos_list, 5) 

    page_number = request.GET.get('page')
    mantenimientos = paginator.get_page(page_number) 
    return render(request, 'flotas/mantenimientos/listar_mantenimientos.html', {'mantenimientos': mantenimientos})


# Crear mantenimiento
def crear_mantenimiento(request):
    if request.method == 'POST':
        form = MantenimientoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_mantenimientos')
    else:
        form = MantenimientoForm()
    return render(request, 'flotas/mantenimientos/crear_mantenimiento.html', {'form': form})

# Editar mantenimiento
def editar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(Mantenimiento, id=id)
    if request.method == 'POST':
        form = MantenimientoForm(request.POST, instance=mantenimiento)
        if form.is_valid():
            form.save()
            return redirect('listar_mantenimientos')
    else:
        form = MantenimientoForm(instance=mantenimiento)
    return render(request, 'flotas/mantenimientos/editar_mantenimiento.html', {'form': form})

# Eliminar mantenimiento
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(Mantenimiento, id=id)
    if request.method == 'POST':
        mantenimiento.delete()
        return redirect('listar_mantenimientos')
    return render(request, 'flotas/mantenimientos/eliminar_mantenimiento.html', {'mantenimiento': mantenimiento})


def listar_alertas(request):
    generar_alertas()  # genera o actualiza las alertas
    alertas_list = AlertaMantenimiento.objects.all().order_by('-fecha_alerta')
    paginator = Paginator(alertas_list, 5)

    page_number = request.GET.get('page')
    alertas = paginator.get_page(page_number)     
    return render(request, 'flotas/alertas/listar_alertas.html', {'alertas': alertas})


def editar_alerta(request, alerta_id):
    alerta = get_object_or_404(AlertaMantenimiento, pk=alerta_id)
    if request.method == 'POST':
        form = AlertaForm(request.POST, instance=alerta)
        if form.is_valid():
            form.save()
            return redirect('listar_alertas')
    else:
        form = AlertaForm(instance=alerta)
    return render(request, 'flotas/alertas/editar_alerta.html', {'form': form})


def eliminar_alerta(request, alerta_id):
    alerta = get_object_or_404(AlertaMantenimiento, id=alerta_id)
    if request.method == 'POST':
        alerta.delete()
    return redirect('listar_alertas')


def historial_mantenimiento(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    mantenimientos = Mantenimiento.objects.filter(id_vehiculo=vehiculo).select_related('id_vehiculo').prefetch_related('detalles__servicio').order_by('-date')
    
    return render(request, 'flotas/historial/historial_mantenimiento.html', {
        'vehiculo': vehiculo,
        'mantenimientos': mantenimientos
    })


def exportar_excel(request):
    vehiculos = Vehiculo.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehículos"

    # Encabezados
    ws.append(['Placa', 'Marca', 'Modelo', 'Año', 'Tipo de Vehículo', 'Estado', 'Kilómetros', 'Último Mto.', 'Próximo Mto.'])

    for v in vehiculos:
        ws.append([
            v.placa,
            v.marca,
            v.modelo,
            v.year,
            v.id_tipo_vehiculo.name if v.id_tipo_vehiculo else '',
            v.get_state_display() if hasattr(v, 'get_state_display') else v.state,
            f"{v.current_kilometers:.1f} km" if v.current_kilometers is not None else '',
            v.last_date_support.strftime("%d/%m/%Y") if v.last_date_support else '',
            v.next_date_support.strftime("%d/%m/%Y") if v.next_date_support else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=vehiculos.xlsx'

    wb.save(response)
    return response


def exportar_excel_con(request):
    conductores = Conductor.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Conductores"

    # Encabezados
    ws.append(['Nombre', 'Apellido', 'DNI', 'Teléfono', 'Dirección', 'Licencia', 'Fecha Ingreso', 'Estado'])

    for c in conductores:
        ws.append([
            c.name,
            c.lastname,
            c.doc_identity,
            c.phone,
            c.address,
            c.licence_drive,
            c.date_entry.strftime("%d/%m/%Y") if c.date_entry else '',
            c.get_state_display() if hasattr(c, 'get_state_display') else c.state,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=conductores.xlsx'

    wb.save(response)
    return response

def exportar_excel_man(request):
    mantenimientos = Mantenimiento.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Mantenimientos"

    # Encabezados según tabla en tu plantilla
    ws.append([
        'Tipo',
        'Vehículo (Placa)',
        'Descripción',
        'Fecha',
        'Kilometraje',
        'Costo (S/)',
        'Taller',
        'Fecha Programada',
        'KM Programado'
    ])

    for m in mantenimientos:
        ws.append([
            m.get_tipo_display() if hasattr(m, 'get_tipo_display') else m.tipo,
            m.id_vehiculo.placa if m.id_vehiculo else '',
            m.description,
            m.date.strftime("%d/%m/%Y") if m.date else '',
            f"{m.kilometraje:.1f} km" if m.kilometraje else '',
            f"S/ {m.costo:.2f}" if m.costo else '',
            m.workshop,
            m.fecha_programada.strftime("%d/%m/%Y") if m.fecha_programada else '—',
            f"{m.km_programado:.1f} km" if m.km_programado else '—',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=mantenimientos.xlsx'

    wb.save(response)
    return response


class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

class TipoVehiculoViewSet(viewsets.ModelViewSet):
    queryset = TipoVehiculo.objects.all()
    serializer_class = TipoVehiculoSerializer

#VEHICULOS
class VehiculoViewSet(viewsets.ModelViewSet):
    queryset = Vehiculo.objects.all()
    serializer_class = VehiculoSerializer
#CONDUCTOR
class ConductorViewSet(viewsets.ModelViewSet):
    queryset = Conductor.objects.all()
    serializer_class = ConductorSerializer

class AsignacionViewSet(viewsets.ModelViewSet):
    queryset = Asignacion.objects.all()
    serializer_class = AsignacionSerializer

class TipoMantenimientoViewSet(viewsets.ModelViewSet):
    queryset = TipoMantenimiento.objects.all()
    serializer_class = TipoMantenimientoSerializer

class MantenimientoViewSet(viewsets.ModelViewSet):
    queryset = Mantenimiento.objects.all()
    serializer_class = MantenimientoSerializer

class AlertaMantenimientoViewSet(viewsets.ModelViewSet):
    queryset = AlertaMantenimiento.objects.all()
    serializer_class = AlertaMantenimientoSerializer

class ServicioMantenimientoViewSet(viewsets.ModelViewSet):
    queryset = ServicioMantenimiento.objects.all()
    serializer_class = ServicioMantenimientoSerializer

class DetalleMantenimientoViewSet(viewsets.ModelViewSet):
    queryset = DetalleMantenimiento.objects.all()
    serializer_class = DetalleMantenimientoSerializer
